import io
import json
import os.path
import time
from datetime import datetime
from typing import List, Annotated
from uuid import UUID
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import Response
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from starlette import status
from uuid6 import uuid7

from app.auth.authenticate import get_current_user
from app.cache import get_cache, set_cache
from app.db.session import get_db
from app.models.user import User
from app.schema.meter import Meter, MeterCreate, MeterUpdate, ResponseModel
from app.schema.user import UserInDB
from app.services.meter import MeterService
from app.utils.convert_meter_to_pydantic import convert_meter_to_pydantic, create_response, meter_to_dict
from app.value_objects.status import StatusState, Status

router = APIRouter()


@router.get("/find_meter_by_user", response_model=ResponseModel, status_code=status.HTTP_200_OK)
async def find_meters_by_user(current_user: Annotated[User, Depends(get_current_user)], 
                              db: AsyncSession = Depends(get_db), 
                              status: StatusState = StatusState.EXECUTING.value):
    service = MeterService(db)
    if status == StatusState.CHECKING.value:
        meters = await service.get_meters_by_user_department(
            department=current_user.department,
            status=status, 
            supervisor=current_user.username
        )
    else:
        meters = await service.get_meters_by_user_department(current_user.department, status)
    data = convert_meter_to_pydantic(meters)
    total = len(data)
    pagination = {
        "offset": 0,
        "limit": total,
        "total": total,
        "order": "asc"
    }
    response = create_response(status_code=200, data=data, pagination=pagination)
    return response


@router.get("/read", response_model=ResponseModel, status_code=status.HTTP_200_OK)
async def read_meters(
        offset: int = Query(0, ge=0, title="Offset", description="Offset for pagination"),
        limit: int = Query(10, ge=10, title="limit", description="Limit for pagination"),
        order: str = Query("asc", regex="^(asc|desc)$", title="Order", description="Order by completion date"),
        db: AsyncSession = Depends(get_db)
):
    service = MeterService(db)
    meters, total = await service.read_meters(offset=offset, limit=limit, order=order)
    data = convert_meter_to_pydantic(meters)
    pagination = {
        'offset': offset,
        'limit': limit,
        'total': total,
        'order': order
    }
    response = {
        "status": 200,
        "result": {
            "data": data,
            "pagination": pagination
        }
    }
    return response


# @router.get("/", response_model=ResponseModel, status_code=status.HTTP_200_OK)
# async def get_meters(
#         status_filter: StatusState = StatusState.EXECUTING.value,
#         db: AsyncSession = Depends(get_db)):
#     service = MeterService(db)
#     meters = await service.get_meters(status=status_filter)
#     data = convert_meter_to_pydantic(meters)
#     total = len(data)
#     pagination = {
#         'offset': 0,
#         'limit': total,
#         'total': total,
#         'order': 'asc'
#     }
#     response = create_response(status_code=200, data=data, pagination=pagination)
#     return response

@router.get("/meters/by-status", response_model=List[Meter])
async def find_all_meters_by_status(status: StatusState, db: AsyncSession = Depends(get_db)):
    service = MeterService(db)
    try:
        meters = await service.find_all_meters_by_status(status.value)
        return meters
    except HTTPException as e:
        raise e

@router.get('/completed', response_model=ResponseModel, status_code=status.HTTP_200_OK)
async def get_completed_meters(db: AsyncSession = Depends(get_db)):
    service = MeterService(db)
    meters = await service.find_completed_meters()
    data = convert_meter_to_pydantic(meters)
    total = len(data)
    pagination = {
        'offset': 0,
        'limit': total,
        'total': total,
        'order': 'asc'
    }
    response = create_response(status_code=200, data=data, pagination=pagination)
    return response


@router.get('/supervisor', response_model=ResponseModel, status_code=status.HTTP_200_OK)
async def get_meters_by_supervisor(current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    service = MeterService(db)
    supervisor = current_user.username
    meters = await service.get_completed_meters_by_supervisor(supervisor)
    data = convert_meter_to_pydantic(meters)
    total = len(data)
    pagination = {
        'offset': 0,
        'limit': 1,
        'total': total,
        'order': 'asc'
    }
    return create_response(status_code=200, data=data, pagination=pagination)


@router.get("/{meter_id}", response_model=ResponseModel, status_code=status.HTTP_200_OK)
async def get_meter(meter_id: UUID, db: AsyncSession = Depends(get_db)):
    service = MeterService(db)
    meter = await service.get_meter(meter_id)
    if not meter:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Meter does not found"
        )
    pagination = {
        'offset': 0,
        'limit': 1,
        'total': 1,
        'order': 'asc'
    }
    data = [meter]
    return create_response(status_code=200, data=data, pagination=pagination)


@router.post("/create", response_model=Meter, status_code=status.HTTP_201_CREATED)
async def create_meter(meter: MeterCreate, db: AsyncSession = Depends(get_db)):
    service = MeterService(db)
    created_meter = await service.create(meter)
    return created_meter


@router.post("/upload", response_model=ResponseModel, status_code=status.HTTP_201_CREATED)
async def upload(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    service = MeterService(db)
    content = file.file.read()
    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook.active
    meters = []

    for row in range(3, sheet.max_row + 1):
        try:
            meter_data = MeterCreate(
                meter_id=uuid7(),
                code=str(sheet.cell(row=row, column=1).value) if sheet.cell(row=row, column=1).value else None,
                owner_name=str(sheet.cell(row=row, column=3).value) if sheet.cell(row=row, column=3) else None,
                meter_number=str(sheet.cell(row=row, column=5).value) if sheet.cell(row=row, column=5) else None,
                address=str(sheet.cell(row=row, column=2).value) if sheet.cell(row=row, column=5) else None,
                previous_reading=sheet.cell(row=row, column=6).value if sheet.cell(row=row, column=6) else None,
                current_reading=sheet.cell(row=row, column=7).value if sheet.cell(row=row, column=6) else None,
                latitude=None,
                longitude=None,
                supervisor=None,
                comment=None,
                completion_date=None,
                photo1_url=None,
                photo2_url=None,
                status=StatusState.EXECUTING
            )
        except KeyError as e:
            raise HTTPException(status_code=400, detail=f"Missing column in the Excel file: {e}")
        meter = await service.create(meter_data)
        meters.append(meter)
    total = len(meters)
    data = convert_meter_to_pydantic(meters)
    pagination = {
        'offset': 0,
        'limit': total,
        'total': total,
        'order': 'asc'
    }
    return create_response(status_code=201, data=data, pagination=pagination)


@router.put("/update/{meter_id}", response_model=Meter, status_code=status.HTTP_200_OK)
async def update_meter(meter_id: UUID, meter: MeterUpdate, db: AsyncSession = Depends(get_db), user: UserInDB = Depends(get_current_user)):
    service = MeterService(db)
    updated_meter = await service.update(meter_id, meter, user)
    if not updated_meter:
        raise HTTPException(status_code=404, detail="Meter not found")
    return updated_meter


@router.delete("/delete_all", response_model=dict, status_code=status.HTTP_200_OK)
async def delete_meters_and_file(filename: str, db: AsyncSession = Depends(get_db)):
    service = MeterService(db)
    await service.delete_all_meters()

    #delete file
    file_path = f"/data/{filename}"
    if os.path.exists(file_path):
        os.remove(file_path)
        return {"detail": "All meters and the file have been deleted."}
    else:
        raise HTTPException(status_code=404, detail="file not found")


@router.delete("/delete", response_model=dict)
async def delete_meters(db: AsyncSession = Depends(get_db)):
    service = MeterService(db)
    await service.delete_all_meters()
    return {"message": "Meters successfully deleted"}


@router.post("/download")
async def download(db: AsyncSession = Depends(get_db)):
    service = MeterService(db)
    file_stream = await service.get_completed_meters()
    file_content = file_stream.getvalue()
    headers = {
        'Content-Disposition': 'attachment; filename="Meters.xlsx"',
        "Access-Control-Allow-Origin": "*",
        "Access-Control-Allow-Headers": "*",
        "Access-Control_Allow-Methods": "POST, GET, OPTIONS",
    }
    return Response(content=file_content,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;charset=utf-8",
                    headers=headers)
